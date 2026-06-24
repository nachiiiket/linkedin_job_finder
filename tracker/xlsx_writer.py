from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL  = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT  = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT    = Font(name="Calibri", size=10)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
THIN_BORDER  = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

COLUMN_CONFIG = [
    ("job_id",           12),
    ("date_found",       16),
    ("position",         30),
    ("company",          22),
    ("location",         18),
    ("poster_name",      20),
    ("poster_title",     28),
    ("poster_profile_url", 35),
    ("job_url",          35),
    ("email",            30),
    ("applied",          10),
    ("connection_sent",  14),
    ("notes",            40),
    ("full_post",        60),
]


def _build_sheet(ws, csv_path: str):
    import csv
    col_map = {}
    for i, (name, width) in enumerate(COLUMN_CONFIG, 1):
        col_map[name] = i
        cell = ws.cell(row=1, column=i, value=name.replace("_", " ").title())
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = width
    ws.row_dimensions[1].height = 22

    with open(csv_path, "r", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row_idx, row in enumerate(reader, 2):
            for field_name, col_idx in col_map.items():
                val = row.get(field_name, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.font = BODY_FONT
                cell.alignment = WRAP_ALIGN
                cell.border = THIN_BORDER

    last_col = get_column_letter(len(COLUMN_CONFIG))
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"
    ws.freeze_panes = "A2"


def export_csv_to_xlsx(csv_path: str, xlsx_path: str = None) -> str:
    csv_path = Path(csv_path)
    if not csv_path.exists():
        return ""

    if xlsx_path is None:
        xlsx_path = csv_path.with_suffix(".xlsx")

    wb = Workbook()
    ws = wb.active
    ws.title = "Jobs"
    _build_sheet(ws, str(csv_path))
    wb.save(str(xlsx_path))

    with open(csv_path, "r", encoding="utf-8") as f:
        import csv
        row_count = sum(1 for _ in csv.DictReader(f))
    print(f"  >> Exported: {xlsx_path} ({row_count} rows)")
    return str(xlsx_path)
